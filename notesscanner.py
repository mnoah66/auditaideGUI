import tkinter as tk
from tkinter import Scrollbar
from tkinter.font import Font
from tkinter import filedialog,messagebox
import openpyxl
from openpyxl import load_workbook
from datetime import timedelta, date, time
import datetime
import csv
import io
import os 
from tkinter import ttk
import xlsxwriter
from checknotesfunctions import flaggedWords, flaggedWordsInverse, oddDuration, oddTimes, underUnits, shortNote, overlapping_notes


class SampleApp(tk.Tk):
    def __init__(self):
        tk.Tk.__init__(self)
        self.title("NotesScanner")

        #scrollbar = Scrollbar(self)
        #scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        nb = ttk.Notebook(self)
        page1 = ttk.Frame(nb)
        page2 = ttk.Frame(nb)


        nb.add(page1, text='Menu')
        nb.add(page2, text='Instructions')

        nb.pack(expand=1, fill="both",padx=10,pady=10)

        


        self.keywordsFrame = tk.LabelFrame(page1,borderwidth=0.5,relief="groove",text="WORDS/PHRASES",font=Font(family='Arial Bold', size=11))
        self.keywordsFrame.pack(padx=20, pady=10,anchor="w",fill=tk.X)

        # -------------------- MAIN GUI -----------------------------
        
        
        self.labelKeywords = tk.Label(self.keywordsFrame, text="Flag notes containing these words/phrases:",font=Font(family='Arial', size=10))
        self.labelKeywords.grid(row=0,column=1,padx=2,pady=2)
        self.varck = tk.IntVar()
        self.checkboxKeywordsInverse = tk.Checkbutton(self.keywordsFrame, variable=self.varck, text="Inverse",command=self.cb).grid(row=0,column=2)
        
        

        self.var1 = tk.IntVar()
        self.var1.set(1)
        self.checkboxKeywords = tk.Checkbutton(self.keywordsFrame, text="Enable", variable=self.var1).grid(row=0,column=0)
        self.entryKeywords = tk.Text(self.keywordsFrame,font=Font(family='Arial', size=10),height=4,width=30)
        self.entryKeywords.grid(row=0,column=3, padx=4,pady=2)
       

        self.durationsGreaterFrame = tk.LabelFrame(page1,borderwidth=1,relief="groove",text="DURATION",font=Font(family='Arial Bold', size=11))
        self.durationsGreaterFrame.pack(padx=20, pady=10,anchor="w",fill=tk.X)
        self.var2 = tk.IntVar()
        self.var2.set(1)
        self.checkboxDurationsGreater = tk.Checkbutton(self.durationsGreaterFrame, text="Enable", variable=self.var2).grid(row=0,column=0)
        self.labelDurationsGreater = tk.Label(self.durationsGreaterFrame, text="Flag notes of duration (minutes) greater/equal to:",font=Font(family='Arial', size=10))
        self.labelDurationsGreater.grid(row=0,column=1,padx=2,pady=2)
        self.spinDurationsGreater = tk.Entry(self.durationsGreaterFrame, width=5,  font=Font(family='Helvetica', size=12))
        self.spinDurationsGreater.grid(row=0,column=2,padx=2,pady=2)
        
    
        self.labelDurationsLess = tk.Label(self.durationsGreaterFrame, text="Flag notes of duration (minutes) less/equal to:",font=Font(family='Arial', size=10))
        self.labelDurationsLess.grid(row=2,column=1,padx=2,pady=2)
        self.spinDurationsLess = tk.Entry(self.durationsGreaterFrame, width=5,  font=Font(family='Helvetica', size=12))
        self.spinDurationsLess.grid(row=2,column=2,padx=2,pady=2)

        
        self.notelengthFrame = tk.LabelFrame(page1,borderwidth=1,relief="groove",text="NOTE LENGTH",font=Font(family='Arial Bold', size=11))
        self.notelengthFrame.pack(padx=20, pady=10,anchor="w",fill=tk.X)
        self.var3 = tk.IntVar()
        self.var3.set(1)
        self.checkboxNoteLength = tk.Checkbutton(self.notelengthFrame, text="Enable", variable=self.var3).grid(row=3,column=0)
        self.labelNoteLength = tk.Label(self.notelengthFrame, text="Flag notes of length under:",font=Font(family='Arial', size=10))
        self.labelNoteLength.grid(row=3,column=1,padx=2,pady=2)
        self.spinNoteLength = tk.Entry(self.notelengthFrame, width=5,  font=Font(family='Helvetica', size=12))
        self.spinNoteLength.grid(row=3,column=2,padx=2,pady=2)

        #### START TIMES
        self.startimeafterFrame = tk.LabelFrame(page1,borderwidth=1,relief="groove",text="START TIME",font=Font(family='Arial Bold', size=11))
        self.startimeafterFrame.pack(padx=20, pady=10,anchor="w",fill=tk.X)
        self.var4 = tk.IntVar()
        self.var4.set(1)
        self.checkboxtimeafter = tk.Checkbutton(self.startimeafterFrame, text="Enable", variable=self.var4).grid(row=4,column=0)
        self.labelStartAfter = tk.Label(self.startimeafterFrame, text="Flag notes with start time after:",font=Font(family='Arial', size=10))
        self.labelStartAfter.grid(row=4,column=1,padx=2,pady=2)
        self.spinHourAfter = tk.Spinbox(self.startimeafterFrame, values=("","01","02","03","04","05","06","07","08","09","10","11","12"),font=Font(family='Helvetica', size=10), width=5,readonlybackground='white')
        self.spinHourAfter.grid(row=4,column=2,padx=2,pady=2)
        self.spinMinAfter = tk.Spinbox(self.startimeafterFrame, values=("","01",
        "02",
        "03",
        "04",
        "05",
        "06",
        "07",
        "08",
        "09",
        "10",
        "11",
        "12",
        "13",
        "14",
        "15",
        "16",
        "17",
        "18",
        "19",
        "20",
        "21",
        "22",
        "23",
        "24",
        "25",
        "26",
        "27",
        "28",
        "29",
        "30",
        "31",
        "32",
        "33",
        "34",
        "35",
        "36",
        "37",
        "38",
        "39",
        "40",
        "41",
        "42",
        "43",
        "44",
        "45",
        "46",
        "47",
        "48",
        "49",
        "50",
        "51",
        "52",
        "53",
        "54",
        "55",
        "56",
        "57",
        "58",
        "59",
        ), font=Font(family='Helvetica', size=10), width=5,readonlybackground='white')
        self.spinMinAfter.grid(row=4,column=3,padx=2,pady=2)
        self.spinAMPMafter = tk.Spinbox(self.startimeafterFrame, values=("","AM","PM"), font=Font(family='Helvetica', size=10), width=5)
        self.spinAMPMafter.grid(row=4,column=4,padx=2,pady=2)

        
        self.labelStartBefore = tk.Label(self.startimeafterFrame, text="Flag notes with start time before:",font=Font(family='Arial', size=10))
        self.labelStartBefore.grid(row=5,column=1,padx=2,pady=2)
        self.spinHourBefore = tk.Spinbox(self.startimeafterFrame, values=("","01","02","03","04","05","06","07","08","09","10","11","12"), font=Font(family='Helvetica', size=10), width=5)
        self.spinHourBefore.grid(row=5,column=2)
        self.spinMinBefore = tk.Spinbox(self.startimeafterFrame, values=("","01",
        "02",
        "03",
        "04",
        "05",
        "06",
        "07",
        "08",
        "09",
        "10",
        "11",
        "12",
        "13",
        "14",
        "15",
        "16",
        "17",
        "18",
        "19",
        "20",
        "21",
        "22",
        "23",
        "24",
        "25",
        "26",
        "27",
        "28",
        "29",
        "30",
        "31",
        "32",
        "33",
        "34",
        "35",
        "36",
        "37",
        "38",
        "39",
        "40",
        "41",
        "42",
        "43",
        "44",
        "45",
        "46",
        "47",
        "48",
        "49",
        "50",
        "51",
        "52",
        "53",
        "54",
        "55",
        "56",
        "57",
        "58",
        "59",
        ), font=Font(family='Helvetica', size=10), width=5)
        self.spinMinBefore.grid(row=5,column=3,padx=2,pady=2)
        self.spinAMPMbefore = tk.Spinbox(self.startimeafterFrame, values=("","AM","PM"), font=Font(family='Helvetica', size=10), width=5)
        self.spinAMPMbefore.grid(row=5,column=4,padx=2,pady=2)

        #### end TIMES
        self.endtimeafterFrame = tk.LabelFrame(page1,borderwidth=1,relief="groove",text="end TIME",font=Font(family='Arial Bold', size=11))
        self.endtimeafterFrame.pack(padx=20, pady=10,anchor="w",fill=tk.X)
        self.var4 = tk.IntVar()
        self.var4.set(1)
        self.checkboxtimeafter = tk.Checkbutton(self.endtimeafterFrame, text="Enable", variable=self.var4).grid(row=4,column=0)
        self.labelendAfter = tk.Label(self.endtimeafterFrame, text="Flag notes with end time after:",font=Font(family='Arial', size=10))
        self.labelendAfter.grid(row=4,column=1,padx=2,pady=2)
        self.endspinHourAfter = tk.Spinbox(self.endtimeafterFrame, values=("","01","02","03","04","05","06","07","08","09","10","11","12"),font=Font(family='Helvetica', size=10), width=5,readonlybackground='white')
        self.endspinHourAfter.grid(row=4,column=2,padx=2,pady=2)
        self.endspinMinAfter = tk.Spinbox(self.endtimeafterFrame, values=("","01",
        "02",
        "03",
        "04",
        "05",
        "06",
        "07",
        "08",
        "09",
        "10",
        "11",
        "12",
        "13",
        "14",
        "15",
        "16",
        "17",
        "18",
        "19",
        "20",
        "21",
        "22",
        "23",
        "24",
        "25",
        "26",
        "27",
        "28",
        "29",
        "30",
        "31",
        "32",
        "33",
        "34",
        "35",
        "36",
        "37",
        "38",
        "39",
        "40",
        "41",
        "42",
        "43",
        "44",
        "45",
        "46",
        "47",
        "48",
        "49",
        "50",
        "51",
        "52",
        "53",
        "54",
        "55",
        "56",
        "57",
        "58",
        "59",
        ), font=Font(family='Helvetica', size=10), width=5,readonlybackground='white')
        self.endspinMinAfter.grid(row=4,column=3,padx=2,pady=2)
        self.endspinAMPMafter = tk.Spinbox(self.endtimeafterFrame, values=("","AM","PM"), font=Font(family='Helvetica', size=10), width=5)
        self.endspinAMPMafter.grid(row=4,column=4,padx=2,pady=2)

        
        self.endlabelendBefore = tk.Label(self.endtimeafterFrame, text="Flag notes with End time before:",font=Font(family='Arial', size=10))
        self.endlabelendBefore.grid(row=5,column=1,padx=2,pady=2)
        self.endspinHourBefore = tk.Spinbox(self.endtimeafterFrame, values=("","01","02","03","04","05","06","07","08","09","10","11","12"), font=Font(family='Helvetica', size=10), width=5)
        self.endspinHourBefore.grid(row=5,column=2)
        self.endspinMinBefore = tk.Spinbox(self.endtimeafterFrame, values=("","01",
        "02",
        "03",
        "04",
        "05",
        "06",
        "07",
        "08",
        "09",
        "10",
        "11",
        "12",
        "13",
        "14",
        "15",
        "16",
        "17",
        "18",
        "19",
        "20",
        "21",
        "22",
        "23",
        "24",
        "25",
        "26",
        "27",
        "28",
        "29",
        "30",
        "31",
        "32",
        "33",
        "34",
        "35",
        "36",
        "37",
        "38",
        "39",
        "40",
        "41",
        "42",
        "43",
        "44",
        "45",
        "46",
        "47",
        "48",
        "49",
        "50",
        "51",
        "52",
        "53",
        "54",
        "55",
        "56",
        "57",
        "58",
        "59",
        ), font=Font(family='Helvetica', size=10), width=5)
        self.endspinMinBefore.grid(row=5,column=3,padx=2,pady=2)
        self.endspinAMPMbefore = tk.Spinbox(self.endtimeafterFrame, values=("","AM","PM"), font=Font(family='Helvetica', size=10), width=5)
        self.endspinAMPMbefore.grid(row=5,column=4,padx=2,pady=2)

        reportsFrame = ttk.Frame(page1)
        reportsFrame.pack(fill=tk.X)
        
        self.unitsFrame = tk.LabelFrame(reportsFrame,borderwidth=1,relief="groove",text="UNDER UNITS",font=Font(family='Arial Bold', size=11))
        self.unitsFrame.pack(side=tk.LEFT, padx=2, pady=10,anchor="w",fill=tk.X)
        self.var5 = tk.IntVar()
        self.var5.set(1)
        self.checkboxunits = tk.Checkbutton(self.unitsFrame, text="Enable", variable=self.var5).grid(row=6,column=0)
        self.labelUnderUnits = tk.Label(self.unitsFrame, text="Flag individuals with total units less than:",font=Font(family='Arial', size=10))
        self.labelUnderUnits.grid(row=6,column=1,padx=2,pady=3)
        self.spinUnderUnits = tk.Entry(self.unitsFrame, width=5,font=Font(family='Helvetica', size=12))
        self.spinUnderUnits.grid(row=6, column=2, padx=2,pady=3)

        self.duptimeFrame = tk.LabelFrame(reportsFrame,borderwidth=1,relief="groove",text="DUPLICATED START TIMES",font=Font(family='Arial Bold', size=11))
        self.duptimeFrame.pack(side=tk.LEFT,padx=2, pady=10,anchor="w",fill=tk.X)
        self.var6 = tk.IntVar()
        self.var6.set(1)
        self.checkboxduptime = tk.Checkbutton(self.duptimeFrame, text="Enable", variable=self.var6).grid(row=7,column=0)

        self.dupcontentFrame = tk.LabelFrame(reportsFrame,borderwidth=1,relief="groove",text="DUPLICATED CONTENT",font=Font(family='Arial Bold', size=11))
        self.dupcontentFrame.pack(side=tk.LEFT,padx=2, pady=10,anchor="w",fill=tk.X)
        self.var6 = tk.IntVar()
        self.var6.set(1)
        self.checkboxdupcontent = tk.Checkbutton(self.dupcontentFrame, text="Enable", variable=self.var6).grid(row=7,column=0)

        
        ### OVERLAPPING SERVICES
        self.overlappingFrame = tk.LabelFrame(reportsFrame,borderwidth=1,relief="groove",text="OVERLAPPING SERVICES",font=Font(family='Arial Bold', size=11))
        self.overlappingFrame.pack(side=tk.LEFT,padx=2, pady=10,anchor="w",fill=tk.X)
        self.var6 = tk.IntVar()
        self.var6.set(1)
        self.checkboxoverlapping = tk.Checkbutton(self.overlappingFrame, text="Enable", variable=self.var6).grid(row=7,column=0)

        ### NON-CONTIGUOUS/GAPS
        self.gapsFrame = tk.LabelFrame(reportsFrame,borderwidth=1,relief="groove",text="NON-CONTIGUOUS/GAPS",font=Font(family='Arial Bold', size=11))
        self.gapsFrame.pack(side=tk.LEFT,padx=2, pady=10,anchor="w",fill=tk.X)
        self.var6 = tk.IntVar()
        self.var6.set(1)
        self.checkboxgaps = tk.Checkbutton(self.gapsFrame, text="Enable", variable=self.var6).grid(row=7,column=0)
 
        ## Choose file to read
        self.buttonFile = tk.Button(page1, text="1. Choose File to be scanned", command=self.file_choose, fg='blue')
        self.buttonFile.pack(fill=tk.X,padx=50,pady=2,anchor="w")
        self.labelFile = tk.Label(page1, text="")
        self.labelFile.pack() 

        # Choose direcgtory to save file ot
        self.buttonFileOutput = tk.Button(page1, text="2. Confirm output file location", command=self.folder_choose, fg='blue')
        self.buttonFileOutput.pack(fill=tk.X,padx=50,pady=2,anchor="w")
        self.labelFileOutput = tk.Entry(page1, text="",background='grey94')
        self.labelFileOutput.pack(fill=tk.X,padx=50,pady=2,anchor="w") 
        
        self.button = tk.Button(page1, text="3. RUN", command=self.on_button, fg='blue')
        self.button.pack(fill=tk.X,padx=50,pady=2,anchor="w") 
        

        self.labelWorking = tk.Label(page1, text="")
        self.labelWorking.pack()


        self.instructions = tk.Text(page2,font=Font(family='Arial', size=10))
        self.instructions.pack(fill=tk.BOTH, expand=1)



        import configparser
        global config 
        config = configparser.ConfigParser()
        config.read('config.ini')
        self.entryKeywords.insert(1.0, config.get('DEFAULT', 'entryKeywords'))
        self.spinDurationsGreater.insert(0, config.get('DEFAULT', 'spinDurationsGreater'))
        self.spinDurationsLess.insert(0, config.get('DEFAULT', 'spinDurationsLess'))
        self.spinNoteLength.insert(0, config.get('DEFAULT', 'spinNoteLength'))
        self.spinHourAfter.insert(0, config.get('DEFAULT', 'spinHourAfter'))
        self.spinMinAfter.insert(0, config.get('DEFAULT', 'spinMinAfter'))
        self.spinAMPMafter.insert(0, config.get('DEFAULT', 'spinAMPMafter'))
        self.spinHourBefore.insert(0, config.get('DEFAULT', 'spinHourBefore'))
        self.spinMinBefore.insert(0, config.get('DEFAULT', 'spinMinBefore'))
        self.spinAMPMbefore.insert(0, config.get('DEFAULT', 'spinAMPMbefore'))
        self.spinUnderUnits.insert(0,config.get('DEFAULT','spinUnderUnits'))
        self.labelFileOutput.insert(0,config.get('DEFAULT','labelFileOutput'))
        self.instructions.insert(1.0, config.get('DEFAULT','instructions'))
        	
        self.link = tk.Label(self, text="")
        self.link.pack()
    #-------------------------  END - MAIN GUI ---------------------------------------------------

    def cb(self):
        

        if self.varck.get() == 1:
            self.labelKeywords.configure(text='Flag notes NOT containing these words/phrases:')
        else:
            self.labelKeywords.configure(text='Flag notes containing these words/phrases:')
    
    def folder_choose(self):
        global dirname
        dirname = filedialog.askdirectory(parent=self, initialdir="/", title='Please select a directory')
        self.labelFileOutput.delete(0, 'end')
        self.labelFileOutput.insert(0,dirname)
        self.labelFileOutput.pack(fill=tk.X,padx=50)    

    def file_choose(self):
        global file_name
        file_name = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if not file_name.endswith(".xlsx"):
            return tk.messagebox.showerror("Warning - File", "Please choose '.xlsx' files only.")
        
        self.labelFile.configure(text=file_name)
        self.labelFile.pack(fill=tk.X,padx=50)
        
    def excel_writer(self):
        global excel_file_name 
        excel_file_name = str(self.labelFileOutput.get()) + '\CheckMyNotes-Created-' + str(datetime.datetime.now().strftime('%Y%m%d%H%M%S'))+ '.xlsx'
        workbook = xlsxwriter.Workbook(excel_file_name)
        worksheet = workbook.add_worksheet()

        format3 = workbook.add_format({'num_format': 'mm/dd/yy'})
        format7 = workbook.add_format({'num_format': 'hh:mm AM/PM'})

        row = 0
        col = 0

        worksheet.write(row,col, 'Flag')
        worksheet.write(row,col+1, 'Individual')
        worksheet.write(row,col+2, 'Start time')
        worksheet.write(row,col+3, 'End Time')
        worksheet.write(row,col+4, 'Date')
        worksheet.write(row,col+5, 'Note/Excerpt')
        worksheet.write(row,col+6, 'Program')
        worksheet.write(row,col+7, 'Duration')
        worksheet.write(row,col+8, 'Note writer')
        worksheet.write(row,col+9, 'Audit Comments')

        from operator import itemgetter
        alpha_list = sorted(results_list, key=itemgetter(1)) # Sort nested list based on the 1th value (individuals name)

        for item in (alpha_list):
            row += 1
            worksheet.write(row, col, item[0])      # The flagged phrase or word
            worksheet.write(row, col +1, item[1])   # The individual
            worksheet.write(row, col +2, item[2], format7)   # The start time
            worksheet.write(row, col +3, item[3], format7)   # The end time
            worksheet.write(row, col +4, item[4], format3)   # The Date
            worksheet.write(row, col +5, item[5])   # The note/excerpt
            worksheet.write(row, col +6, item[6])   # The program name
            worksheet.write(row, col +7, item[7])   # The Duration
            worksheet.write(row, col +8, item[8])   # The Duration
        workbook.close()

    def on_button(self):        
        # results_list gets passed around from function to function
        # and then written to excel file
        global results_list
        results_list = []

        
        # Update the config.ini file with user settings
        self.saveConfig()
        
        
        #  File validation, then open in memory
        try:
            file_name
            with open(file_name, "rb") as f:
                in_mem_file = io.BytesIO(f.read())
            trngfile = openpyxl.load_workbook(in_mem_file, read_only=True)
            ws = trngfile.active
        except:
            return tk.messagebox.showerror("Warning - File", "An error occurred with the file.  Please choose an .xlsx file only.")

        if self.labelFileOutput.get() == "":
            return  tk.messagebox.showerror("Warning - Directory", "Please choose a save location.") 
        
        # ---------- VARIABLES TO PASS TO FUNCTIONS -----------------
        
        #  Create variables from GUI that will be passed to functions
        keywords = self.entryKeywords.get("1.0", 'end-1c')
        my_list0 = keywords.split(",")
        my_list = [x.strip().lower().replace('?',' ') for x in my_list0]
        print(my_list) 
        try:
            greaterthan = int(self.spinDurationsGreater.get())
            lessthan = int(self.spinDurationsLess.get())
            notelength = int(self.spinNoteLength.get())
            unitThreshold = int(self.spinUnderUnits.get())
        except (TypeError, ValueError):
            return tk.messagebox.showerror("Warning - Integer", "Please enter whole numbers only (e.g. 360 or 12)")
        
        startTimeAfter = self.spinHourAfter.get() + ":" + self.spinMinAfter.get() + " " + self.spinAMPMafter.get()
        startTimeBefore = self.spinHourBefore.get() + ":" + self.spinMinBefore.get() + " " + self.spinAMPMbefore.get()
        
        if self.var1.get() == 1 and self.varck.get() == 1:
            flaggedWordsInverse(ws, my_list, results_list)
        elif self.var1.get() == 1:
            flaggedWords(ws, my_list, results_list)
        if self.var2.get():
            oddDuration(ws,greaterthan,lessthan,results_list)
        if self.var3.get():
            shortNote(ws, notelength,results_list)
        if self.var4.get():
            oddTimes(ws, startTimeAfter,startTimeBefore,results_list)
        if self.var5.get():
            underUnits(ws, unitThreshold,results_list)
        if self.var6.get():
            overlapping_notes(ws, results_list)
        
        def callback(event):
            import os
            import webbrowser
            webbrowser.open_new(r"file://" + os.path.abspath(str(excel_file_name)))
            self.link.configure(text="")
            self.labelWorking.configure(text="")
        
        if len(results_list) != 0:
            self.excel_writer()
            self.labelWorking.configure(font=Font(family='Helvetica', size=12),text="FINISHED!")
            self.link.bind("<Button-1>", callback)
            self.link.configure(text="Click for file", fg="blue", cursor="hand2")
        else:
            tk.messagebox.showerror("No results", "There were no results.  Check your data file and settings and try again.")

        
        
        
        
        self.labelWorking.bind("<Button-1>", callback)


        
    def saveConfig(self):
        config.set('DEFAULT', 'entryKeywords', self.entryKeywords.get("1.0",'end-1c'))
        config.set('DEFAULT', 'spinDurationsGreater', self.spinDurationsGreater.get())
        config.set('DEFAULT', 'spinDurationsLess', self.spinDurationsLess.get())
        config.set('DEFAULT', 'spinNoteLength', self.spinNoteLength.get())
        config.set('DEFAULT', 'spinHourAfter', self.spinHourAfter.get())
        config.set('DEFAULT', 'spinMinAfter', self.spinMinAfter.get())
        config.set('DEFAULT', 'spinAMPMafter', self.spinAMPMafter.get())
        config.set('DEFAULT', 'spinHourBefore', self.spinHourBefore.get())
        config.set('DEFAULT', 'spinMinBefore', self.spinMinBefore.get())
        config.set('DEFAULT', 'spinAMPMbefore', self.spinAMPMbefore.get())
        config.set('DEFAULT', 'spinUnderUnits', self.spinUnderUnits.get())
        config.set('DEFAULT', 'labelFileOutput', self.labelFileOutput.get())
        config.set('DEFAULT', 'instructions', self.instructions.get("1.0",'end-1c'))
        config.write(open('config.ini','w'))
    
        
app = SampleApp()
app.mainloop()